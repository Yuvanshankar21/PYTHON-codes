from openpyxl import Workbook

a = Workbook()

b = a.active

ws = a.create_sheet("first sheet")
column = ['aravind', 'yuvan', 'ajay']

for cols in ws.iter_cols(min_col=2, max_row=1, max_col=2):
    print(cols)
    for row in column:
        ws.append(column)

a.save("myxl.xlsx")




